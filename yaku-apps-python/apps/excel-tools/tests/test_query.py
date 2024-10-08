import pytest
from openpyxl import Workbook
from yaku.excel_tools.commands.query import query


@pytest.fixture
def test():
    return "test_query"


def test_query():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "key"
    sheet["B1"] = "value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"

    query(wb, "key == 'Key2'", "result")
    assert sheet["C1"].value == "result"
    assert sheet["C2"].value == "True"


def test_query_with_sheet_name():
    wb = Workbook()
    wb.create_sheet("Sheet2")
    sheet = wb["Sheet"]
    sheet["A1"] = "key"
    sheet["B1"] = "value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"

    query(wb, "key == 'Key2'", "result", sheet_name="Sheet")
    assert sheet["C1"].value == "result"
    assert sheet["C2"].value == "True"


def test_query_with_true_value():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "key"
    sheet["B1"] = "value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"

    query(wb, "key == 'Key2'", "result", sheet_name="Sheet", true_value="Yes")
    assert sheet["C1"].value == "result"
    assert sheet["C2"].value == "Yes"


def test_query_with_false_value():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "key"
    sheet["B1"] = "value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"

    query(wb, "key != 'Key2'", "result", sheet_name="Sheet", false_value="No")
    assert sheet["C1"].value == "result"
    assert sheet["C2"].value == "No"
