import pytest
from openpyxl import Workbook
from yaku.excel_tools.commands.mapping import apply_list, apply_mapping, get, get_column


@pytest.fixture
def test():
    return "test_mapping"


def test_get():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"

    mapping = get(wb, "Sheet", "A", "B")
    assert mapping == {"Key": "Value", "Key2": "Value2"}


def test_apply_mapping():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"

    mapping = get(wb, "Sheet", "A", "B")
    apply_mapping(wb, mapping, "Mapped")
    assert sheet["C1"].value == "Mapped"
    assert sheet["C2"].value == "Value2"


def test_apply_mapping_does_not_touch_existing_column_headers():
    HEADER = ["ColumnHeader[Key]", "ColumnHeader[Value]", "ColumnHeader[Mapped]"]

    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = HEADER[0]
    sheet["B1"] = HEADER[1]
    sheet["A2"] = "Some Key"
    sheet["B2"] = "Some Value"

    mapping = {HEADER[0]: "unused"}
    apply_mapping(wb, mapping, HEADER[2])

    assert sheet["A1"].value == HEADER[0]
    assert sheet["B1"].value == HEADER[1]
    assert sheet["C1"].value == HEADER[2]


def test_apply_mapping_with_trailing_whitespace():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"

    sheet["A2"] = "MapKey"
    sheet["B2"] = "Some Value"

    sheet["A3"] = "NotInMapKey"
    sheet["B3"] = "Some other value with no mapping"

    sheet["A4"] = "MapKey "  # with a trailing whitespace!
    sheet["B4"] = "Another value"

    mapping = {"MapKey": "MappingValueForKey3"}
    apply_mapping(wb, mapping, "Mapped")

    assert sheet["C2"].value == mapping["MapKey"]
    assert sheet["C3"].value == ""
    assert sheet["C4"].value == mapping["MapKey"]


def test_get_column():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["A2"] = "Key2"

    column = get_column(wb, "Sheet", "A")
    assert column == ["Key", "Key2"]


def test_apply_list_both_set():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "Keys"
    sheet["A2"] = "Key1"
    sheet["A3"] = "Key2"

    apply_list(
        wb,
        ["Key1"],
        "Mapped",
        check_column_names=["A"],
        true_value="true",
        false_value="false",
    )
    assert sheet["B1"].value == "Mapped"
    assert sheet["B2"].value == "true"
    assert sheet["B3"].value == "false"


def test_apply_list_true_set():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "Keys"
    sheet["A2"] = "Key1"
    sheet["A3"] = "Key2"

    apply_list(wb, ["Key1"], "Mapped", check_column_names=["A"], true_value="true")
    assert sheet["B1"].value == "Mapped"
    assert sheet["B2"].value == "true"
    assert sheet["B3"].value == "No"


def test_apply_list_false_set():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "Keys"
    sheet["A2"] = "Key1"
    sheet["A3"] = "Key2"

    apply_list(wb, ["Key1"], "Mapped", check_column_names=["A"], false_value="false")
    assert sheet["B1"].value == "Mapped"
    assert sheet["B2"].value == "Yes"
    assert sheet["B3"].value == "false"


def test_apply_list_column_name():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "KeysA"
    sheet["A2"] = "Key1"
    sheet["A3"] = "Key2"

    sheet["B1"] = "KeysB"
    sheet["B2"] = "Key2"
    sheet["B3"] = "Key1"

    apply_list(wb, ["Key1"], "Mapped", check_column_names=["KeysB"])
    assert sheet["C1"].value == "Mapped"
    assert sheet["C2"].value == "No"
    assert sheet["C3"].value == "Yes"


def test_apply_list_column_name_with_excel_index():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "KeysA"
    sheet["A2"] = "Key1"
    sheet["A3"] = "Key2"

    sheet["B1"] = "KeysB"
    sheet["B2"] = "Key2"
    sheet["B3"] = "Key1"

    apply_list(wb, ["Key1"], "Mapped", check_column_names=["B"])
    assert sheet["C1"].value == "Mapped"
    assert sheet["C2"].value == "No"
    assert sheet["C3"].value == "Yes"


def test_apply_list_column_names():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "KeysA"
    sheet["A2"] = "Key1"
    sheet["A3"] = "Key2"

    sheet["B1"] = "KeysB"
    sheet["B2"] = "Key2"
    sheet["B3"] = "Key1"

    apply_list(wb, ["Key1"], "Mapped", check_column_names=["KeysB", "KeysC"])
    assert sheet["C1"].value == "Mapped"
    assert sheet["C2"].value == "No"
    assert sheet["C3"].value == "Yes"


def test_apply_list_column_names_check_columns_not_found():
    # if the check_column_names are not found, the column will be ignored / default to false
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "KeysA"
    sheet["A2"] = "Key1"
    sheet["A3"] = "Key2"
    sheet["A4"] = "Key3"

    sheet["B1"] = "KeysB"
    sheet["B2"] = "Key2"
    sheet["B3"] = "Key1"
    sheet["B4"] = "Key3"

    apply_list(wb, ["Key1"], "Mapped", check_column_names=["KeysB", "KeysC"])
    assert sheet["C1"].value == "Mapped"
    assert sheet["C2"].value == "No"
    assert sheet["C3"].value == "Yes"
    assert sheet["C4"].value == "No"
