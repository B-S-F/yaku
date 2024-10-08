from pathlib import Path

import openpyxl.worksheet.hyperlink
import pandas as pd
import pytest
from openpyxl import Workbook
from yaku.excel_tools.utils.excel import (
    add_column_to_dataframes,
    add_column_to_sheets,
    clean_column_names,
    get_cell_length,
    get_mapping,
    load_dataframes,
    map_values,
    parse_excel_with_hyperlinks,
    resize_col,
    wrap_row_text,
    write_dataframes,
)


@pytest.fixture
def test():
    return "test_format"


def test_get_mapping():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    mapping = get_mapping(sheet, "A", "B")
    assert mapping == {"Key": "Value"}

    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"
    mapping = get_mapping(sheet, "A", "B")
    assert mapping == {"Key": "Value", "Key2": "Value2"}


def test_map_values():
    keys = ["Key", "Key2"]
    mapping = {"Key": "Value", "Key2": "Value2"}
    values = map_values(keys, mapping)
    assert values == ["Value", "Value2"]

    keys = ["Key", "Key2"]
    mapping = {"Key": "Value"}
    values = map_values(keys, mapping)
    assert values == ["Value", ""]


def test_resize_col():
    wb = Workbook()
    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    initial_width = sheet.column_dimensions["A"].width
    sheet["A3"] = "I am a longer string than the initial width"

    resize_col(sheet, "A")
    second_width = sheet.column_dimensions["A"].width
    assert second_width > initial_width
    sheet["A5"] = "I am a longer string than the longest string in the column"

    resize_col(sheet, "A")
    third_width = sheet.column_dimensions["A"].width
    assert third_width > second_width


def test_resize_col_with_empty_sheet():
    wb = Workbook()
    sheet = wb["Sheet"]
    resize_col(sheet, "A")
    assert sheet.column_dimensions["A"].width > 0.0


def test_get_cell_length():
    class Cell:
        def __init__(self, value):
            self.value = value

    assert get_cell_length(Cell("test")) == 4.0


def test_get_cell_length_with_Falsy_value():
    class Cell:
        def __init__(self, value):
            self.value = value

    falsy_values = [None, "", 0, False, [], {}]

    for value in falsy_values:
        assert get_cell_length(Cell(value)) == 0.0


def test_wrap_row_text():
    wb = Workbook()

    sheet = wb["Sheet"]
    sheet["A1"] = "some text"
    wrap_row_text(sheet, 1)
    cell = sheet["A1"]
    assert cell.alignment.wrap_text is True
    assert cell.alignment.vertical == "top"


def test_load_dataframes():
    wb = Workbook()

    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"

    dfs = load_dataframes(wb)
    assert not dfs["Sheet"].empty
    assert dfs["Sheet"].columns[0] == "Key"
    assert dfs["Sheet"].columns[1] == "Value"


def test_load_dataframe_with_sheet_name():
    wb = Workbook()

    wb.create_sheet("Sheet1")
    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"

    dfs = load_dataframes(wb, sheet_name="Sheet")
    assert not dfs["Sheet"].empty
    assert dfs["Sheet"].columns[0] == "Key"
    assert dfs["Sheet"].columns[1] == "Value"
    assert "Sheet1" not in dfs


def test_add_column_to_dataframes():
    wb = Workbook()

    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"
    dfs = load_dataframes(wb)

    dfs = add_column_to_dataframes(dfs, "New Column", ["New Value"])
    assert dfs["Sheet"].columns[2] == "New Column"
    assert dfs["Sheet"].iloc[0, 2] == "New Value"


def test_add_column_to_sheets():
    wb = Workbook()

    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"

    add_column_to_sheets(wb, "New Column", ["New Value"])
    assert sheet["C1"].value == "New Column"
    assert sheet["C2"].value == "New Value"


def test_clean_column_names():
    wb = Workbook()

    sheet = wb["Sheet"]
    sheet["A1"] = "Not a clean column name"
    sheet["B1"] = "Not a clean !!column //name-1"

    clean_column_names(wb)
    assert sheet["A1"].value == "not_a_clean_column_name"
    assert sheet["B1"].value == "not_a_clean_column_name_1"


def test_write_dataframes():
    wb = Workbook()

    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"
    dfs = load_dataframes(wb)

    dfs = add_column_to_dataframes(dfs, "New Column", ["New Value"])
    write_dataframes(wb, dfs)
    assert sheet["C1"].value == "New Column"
    assert sheet["C2"].value == "New Value"


def test_write_dataframes_handles_empty_sheet():
    wb = Workbook()

    sheet = wb["Sheet"]
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    sheet["A2"] = "Key2"
    sheet["B2"] = "Value2"
    dfs = load_dataframes(wb)

    dfs = add_column_to_dataframes(dfs, "New Column", ["New Value"])
    wb.create_sheet("Sheet2")
    write_dataframes(wb, dfs)
    assert sheet["C1"].value == "New Column"
    assert sheet["C2"].value == "New Value"
    assert "Sheet2" in wb.sheetnames


@pytest.mark.parametrize(
    ("hyperlink_column", "expected_all_hyperlinks"),
    [(0, False), (1, False), (2, True)],
)
@pytest.mark.parametrize(
    "xlsfile",
    [
        # ("table_with_hyperlinks.xls",),
        "table_with_hyperlinks.xlsx",
    ],
)
def test_parse_excel_with_hyperlinks(xlsfile, hyperlink_column, expected_all_hyperlinks):
    xl = (Path(__file__).parent) / "data" / xlsfile
    df, all_hyperlinks = parse_excel_with_hyperlinks(
        xl, "Tabelle1", [hyperlink_column], header_row=0
    )

    expected_data = {
        "Column A": ["Amir", "Anton", "Aslan"],
        "Column B": ["Basketball", "Boxing", "Busdriving"],
        "Column C": ["Catbird", "Condor", "Crow"],
    }
    for index in range(len(df.columns)):
        column_name = df.columns[index]
        if index == hyperlink_column:
            values = df[column_name]
            for row_index, value in enumerate(values):
                if isinstance(value, openpyxl.worksheet.hyperlink.Hyperlink):
                    ref = value.ref
                    assert (
                        ref[0] == column_name[-1]
                    )  # check column name of cell of the hyperlink
                    assert (
                        int(ref[1]) - 1 == row_index + 1
                    )  # check row number of cell of the hyperlink
                else:
                    assert df[column_name][row_index] == expected_data[column_name][row_index]
        else:
            pd.testing.assert_frame_equal(
                df[[column_name]],
                pd.DataFrame.from_dict({column_name: expected_data[column_name]}),
                check_dtype=False,
            )
    assert all_hyperlinks == expected_all_hyperlinks
