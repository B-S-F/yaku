import openpyxl
import pytest
from yaku.excel_tools.commands.print_cell_value import (
    print_cell_value_from_filtered_row,
)


@pytest.fixture
def make_workbook():
    def _make_workbook(sheet_name, data, hyperlinks=None):
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet(sheet_name)
        for y, row in enumerate(data, start=1):
            for x, value in enumerate(row, start=1):
                cell = sheet.cell(row=y, column=x)
                cell.value = value
                if hyperlinks:
                    if (x, y) in hyperlinks:
                        cell.hyperlink = hyperlinks[(x, y)]
        return wb

    return _make_workbook


def test_print_cell_value_can_filter_data(make_workbook, capsys):
    wb = make_workbook("Sheet 1", [["1st row", "B1 value"], ["2nd row", "B2 value"]])
    print_cell_value_from_filtered_row(
        workbook=wb,
        sheet_name="Sheet 1",
        filter_column="A",
        filter_value="1st row",
        column="B",
    )
    captured = capsys.readouterr()
    assert captured.out == "B1 value\n"

    print_cell_value_from_filtered_row(
        workbook=wb,
        sheet_name="Sheet 1",
        filter_column="A",
        filter_value="2nd row",
        column="B",
    )
    captured = capsys.readouterr()
    assert captured.out == "B2 value\n"


def test_print_cell_value_can_return_multiple_matches(make_workbook, capsys):
    wb = make_workbook("Sheet 1", [["A", "one"], ["B", "two"], ["A", "three"]])

    print_cell_value_from_filtered_row(
        workbook=wb,
        sheet_name="Sheet 1",
        filter_column="A",
        filter_value="A",
        column="B",
        return_all=False,  # first, we only return the first match
    )
    captured = capsys.readouterr()
    assert captured.out == "one\n"

    print_cell_value_from_filtered_row(
        workbook=wb,
        sheet_name="Sheet 1",
        filter_column="A",
        filter_value="A",
        column="B",
        return_all=True,  # now we return all matches
    )
    captured = capsys.readouterr()
    assert captured.out == "one\nthree\n"


def test_print_cell_value_can_skip_rows(make_workbook, capsys):
    wb = make_workbook("Sheet 1", [["A", "one"], ["B", "two"], ["A", "three"]])

    print_cell_value_from_filtered_row(
        workbook=wb,
        sheet_name="Sheet 1",
        filter_column="A",
        filter_value="A",
        column="B",
        return_all=True,  # there are two "A" rows (1st and 3rd)
        skip_rows=1,  # but we skip the first row
    )
    captured = capsys.readouterr()
    assert captured.out == "three\n"


def test_print_cell_value_can_extract_hyperlinks(make_workbook, capsys):
    wb = make_workbook(
        "Sheet 1", [["A", "one"], ["B", "two"]], hyperlinks={(2, 1): "http://some.where/"}
    )

    print_cell_value_from_filtered_row(
        workbook=wb,
        sheet_name="Sheet 1",
        filter_column="A",
        filter_value="A",
        column="B",
        url=True,
    )
    captured = capsys.readouterr()
    assert captured.out == "http://some.where/\n"


def test_print_cell_value_raises_error_if_no_hyperlink(make_workbook):
    wb = make_workbook("Sheet 1", [["A", "one"], ["B", "two"]])

    with pytest.raises(RuntimeError, match="has no hyperlink"):
        print_cell_value_from_filtered_row(
            workbook=wb,
            sheet_name="Sheet 1",
            filter_column="A",
            filter_value="A",
            column="B",
            url=True,
        )
