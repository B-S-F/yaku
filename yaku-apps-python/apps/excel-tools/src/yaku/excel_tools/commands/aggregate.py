from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from pandas import DataFrame, read_csv

from ..utils.excel import write_dataframes


def get_csv(csv_path: Path) -> DataFrame:
    """Return the csv at the path as a list of lists."""
    try:
        return read_csv(csv_path, encoding="utf-8")
    except Exception:
        return DataFrame()


def create_sheets(workbook: Workbook, csvs: list[Path]):
    """Create a sheet in the workbook for each csv in the folder."""
    dfs: dict[str, DataFrame] = {}
    for csv in csvs:
        sheet_name = csv.stem[0:31]  # Excel only supports 31 characters for sheet names
        if len(csv.stem) > 31:
            logger.warning(
                "Have to truncate sheet name `{long}` to `{short}` "
                "because Excel cannot read such long names!",
                long=csv.stem,
                short=sheet_name,
            )
        workbook.create_sheet(sheet_name)
        dfs[sheet_name] = get_csv(csv)
    write_dataframes(workbook, dfs)


def aggregate(csv_folder: Path, output_path, glob_pattern="*.csv"):
    folder_elements = csv_folder.glob(glob_pattern)
    csvs = [x for x in folder_elements if x.is_file()]
    if not csvs:
        raise Exception(f"No files found in {csv_folder}")
    workbook = Workbook()
    # remove the default sheet
    workbook.remove(workbook.get_sheet_by_name("Sheet"))
    create_sheets(workbook, csvs)
    workbook.save(output_path)
