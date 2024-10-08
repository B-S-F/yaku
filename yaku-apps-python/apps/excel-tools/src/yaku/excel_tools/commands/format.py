from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from ..utils.excel import resize_col, wrap_row_text


def check_column_limit(column: str) -> bool:
    max_excel_column = "XFD"
    if len(column) < 3:
        return True
    if len(column) > 3:
        return False
    if column <= max_excel_column:
        return True
    return False


def format(workbook: Workbook, column: str, row: int, sheet_name=None, all=False):
    """Format the xlsx file by resizing the column and wrapping the row."""
    if not check_column_limit(column):
        raise ValueError(f"Maximum column is 'XFD' but column was {column}")
    sheets = workbook.sheetnames
    if sheet_name:
        sheets = [sheet_name]
    for sheet_name in sheets:
        sheet = workbook[sheet_name]
        assert isinstance(sheet, Worksheet)
        if all:
            for col_index in range(0, sheet.max_column):
                col = sheet.cell(row=row, column=col_index + 1).column_letter
                resize_col(sheet, col)
        else:
            resize_col(sheet, column)
        wrap_row_text(sheet, row)
