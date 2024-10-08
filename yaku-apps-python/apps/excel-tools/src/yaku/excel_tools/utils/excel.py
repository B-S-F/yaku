from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook, styles
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from .dataframe import add_column
from .vendored.skimpy import clean_columns


def load_dataframes(workbook: Workbook, sheet_name=None) -> dict[str, pd.DataFrame]:
    """Load all the dataframes from the workbook."""
    dataframes: dict[str, pd.DataFrame] = {}
    worksheets: list
    if sheet_name:
        worksheets = [workbook[sheet_name]]
    else:
        worksheets = workbook.worksheets
    for sheet in worksheets:
        dataframe = pd.read_excel(workbook, sheet_name=sheet.title, engine="openpyxl")
        dataframes[sheet.title] = dataframe
    return dataframes


def write_dataframes(workbook: Workbook, dataframes: dict[str, pd.DataFrame]):
    """Write all the dataframes to the workbook."""
    for sheet in workbook.worksheets:
        assert isinstance(sheet, Worksheet)
        if sheet.title not in dataframes:
            continue
        dataframe = dataframes[sheet.title]
        sheet.delete_rows(1, sheet.max_row)
        for row in dataframe_to_rows(dataframe, index=False, header=True):
            sheet.append(row)


def add_column_to_dataframes(
    dataframes: dict[str, pd.DataFrame], column_name: str, column_values: list
) -> dict[str, pd.DataFrame]:
    """Add a column to each dataframe in the workbook."""
    for dataframe in dataframes.values():
        dataframe = add_column(dataframe, column_name, column_values)
    return dataframes


def add_column_to_sheets(workbook: Workbook, column_name: str, column_values: list):
    """Add a column to each sheet in the workbook."""
    dataframes = load_dataframes(workbook)
    dataframes = add_column_to_dataframes(dataframes, column_name, column_values)
    write_dataframes(workbook, dataframes)


def get_mapping(sheet, key_column: str, value_column: str) -> dict[str, str]:
    """Return a dictionary with the mapping of the key_column to the value_column."""
    mapping = {}
    for rows in sheet.rows:
        key = rows[ord(key_column) - 65].value
        value = rows[ord(value_column) - 65].value
        mapping[key] = value
    return mapping


def get_list(sheet, column: str) -> list[str]:
    """Return a list of the values in the column."""
    column_values = []
    for rows in sheet.rows:
        column_values.append(rows[ord(column) - 65].value)
    return column_values


def map_values(keys: list[str], mapping: dict[str, str]) -> list[str]:
    """Map the values in the keys list to the values in the mapping dictionary."""
    column = []
    for key in keys:
        stripped_key = key.rstrip()
        if stripped_key in mapping:
            column.append(mapping[stripped_key])
        else:
            column.append("")
    return column


def convert_cell_value_to_string(cell) -> str:
    """
    Convert the cell value to a string.

    If the value is not convertible to a string return an empty string.
    """
    try:
        if cell.value:
            return str(cell.value)
    except (ValueError, TypeError):
        return ""
    return ""


def get_cell_length(cell) -> float:
    """Return the length of the cell value."""
    s = convert_cell_value_to_string(cell)
    return float(len(s))


def resize_col(sheet, col_letter: str):
    """Resize the column to fit the longest value in the column."""
    max_length = 0.0
    col = sheet[col_letter]
    for cell in col:
        if cell and cell.value:
            cell_length = get_cell_length(cell)
            if cell_length > max_length:
                max_length = cell_length
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[col_letter].width = adjusted_width


def wrap_row_text(sheet, row_number: int):
    """Wrap the text in the row."""
    row = sheet[row_number]
    for cell in row:
        cell.alignment = styles.Alignment(wrap_text=True, vertical="top")


def clean_column_names(workbook: Workbook):
    """Fix the column names in the xlsx file."""
    dataframes = load_dataframes(workbook)
    for key, dataframe in dataframes.items():
        dataframe = clean_columns(dataframe)
        dataframes[key] = dataframe
    write_dataframes(workbook, dataframes)


def column_letter_to_index(column_letter: str) -> int:
    """Convert an Excel column letter to a 0-indexed column index."""
    index = 0
    for i, letter in enumerate(reversed(column_letter)):
        index += (ord(letter.upper()) - 64) * (26**i)
    return index - 1


def _get_link_if_exists(ws, row: int, column: int) -> tuple[str | None, bool]:
    # openpyxl cell-access is 1-based
    cell = ws.cell(row=row + 1, column=column + 1)
    if cell.hyperlink:
        return (cell.hyperlink, True)
    else:
        return (cell.value, False)


def parse_excel_with_hyperlinks(
    excel_path: Path,
    sheet_name: str | int,
    hyperref_cols: list[int],
    header_row: int,
) -> tuple[pd.DataFrame, bool | None]:
    """
    Read an Excel file and replace cell values by hyperlinks if they are given.

    The Excel file given by `excel_path` is read and its `sheet_name` is opened.
    For each column index in `hyperref_cols` (0-based), the cells' values are
    replaced by the hyperlink of that cell (if there is a hyperlink given).

    The `header_row` (0-based) points to the row with header labels. For example
    if `header_row=0`, the first row in the sheet will be ignored.

    Returns the modified DataFrame and a flag which indicates whether all cells
    in the given columns have a hyperlink or not.
    """
    df = pd.read_excel(excel_path, sheet_name, header=header_row)
    wb = load_workbook(excel_path)
    if isinstance(sheet_name, int):
        sheet_name = wb.sheetnames[sheet_name]
    ws = wb[sheet_name]

    all_hyperlinks = True if hyperref_cols else None
    for column in hyperref_cols:
        column_name = df.columns[column]
        row_offset = header_row + 1
        hyperref_col = []
        for i in range(len(df[column_name])):
            (link, is_hyperlink) = _get_link_if_exists(ws, row=row_offset + i, column=column)
            hyperref_col.append(link)
            if not is_hyperlink:
                all_hyperlinks = False
        df[column_name] = hyperref_col
    return (df, all_hyperlinks)
