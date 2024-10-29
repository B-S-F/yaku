from pathlib import Path

import click
from openpyxl import load_workbook

from .commands.aggregate import aggregate as aggregate_command
from .commands.evaluate.main import check_columns as check_columns_command
from .commands.format import format
from .commands.mapping import apply_list, apply_mapping, get, get_column
from .commands.print_cell_value import print_cell_value_from_filtered_row
from .commands.query import query as query_command
from .utils.excel import add_column_to_sheets


def print_version(ctx, param, value):
    import importlib.resources

    if not value or ctx.resilient_parsing:
        return
    version = importlib.resources.read_text("yaku.excel_tools", "_version.txt")
    click.echo(version.strip())
    ctx.exit()


@click.group(invoke_without_command=True, no_args_is_help=True)
@click.option(
    "--version",
    is_flag=True,
    callback=print_version,
    expose_value=False,
    is_eager=True,
    help="Output version information and exit",
)
def main():
    pass


@main.command()
@click.option(
    "--csv-path",
    default=".",
    help="The path to the csv files to aggregate",
    show_default=True,
    type=click.Path(exists=True),
)
@click.option(
    "--output-path",
    default="aggregate.xlsx",
    help="The output path for the excel file",
    show_default=True,
)
@click.option(
    "--glob-pattern",
    default="*.csv",
    help="The glob pattern to match the csv files",
    show_default=True,
)
def aggregate(csv_path: str, output_path: str, glob_pattern: str):
    csv_folder = Path(csv_path)
    output_file = Path(output_path)
    aggregate_command(csv_folder, output_file, glob_pattern)


@main.command()
@click.option(
    "--xlsx-path",
    default="aggregate.xlsx",
    help="The path of the excel file",
    show_default=True,
)
@click.option(
    "--mapping-path",
    help="The output path for the excel file where the mapping is located",
    required=True,
    type=click.Path(exists=True),
)
@click.option(
    "--sheet-name",
    default="ML",
    help="The name of the sheet get the mapping from",
    show_default=True,
)
@click.option(
    "--key-column",
    default="A",
    help="The name of the column to get the mapping keys from",
    show_default=True,
)
@click.option(
    "--value-column",
    help="The name of the column to get the mapping values from, if not provided will use true_value and false_value",
)
@click.option(
    "--check-column-identifiers",
    help="A list of column names to check for the keys. If the column does not exist, it will be ignored. Can also contain Excel column indexes (e.g. 'A', 'B', 'C', etc.).",
    default="A",
    show_default=True,
)
@click.option(
    "--true-value",
    default="Yes",
    help="The value to set when the key is in the mapping",
    show_default=True,
)
@click.option(
    "--false-value",
    default="No",
    help="The value to set when the key is not in the mapping",
    show_default=True,
)
@click.option(
    "--mapping-column-name",
    default="ML-C",
    help="The name of the column to put the mapping values from",
)
def add_mapping(
    xlsx_path: str,
    mapping_path: str,
    sheet_name: str,
    key_column: str,
    value_column: str,
    check_column_identifiers: list[str],
    true_value: str,
    false_value: str,
    mapping_column_name: str,
):
    mapping_workbook = load_workbook(mapping_path)
    output_workbook = load_workbook(xlsx_path)
    if value_column is None:
        parsed_check_column_identifiers = [
            item.strip() for item in check_column_identifiers.split(",")
        ]
        keys = get_column(mapping_workbook, sheet_name, key_column)
        apply_list(
            output_workbook,
            keys,
            mapping_column_name,
            parsed_check_column_identifiers,
            true_value,
            false_value,
        )
    else:
        mapping = get(mapping_workbook, sheet_name, key_column, value_column)
        apply_mapping(output_workbook, mapping, mapping_column_name)

    output_workbook.save(xlsx_path)


@main.command()
@click.argument("xlsx_path", type=click.Path(exists=True), required=True)
@click.option(
    "--sheet-name",
    help="The name of the sheet to format, if not provided will format all sheets",
)
@click.option(
    "--column",
    default="A",
    help="The column of the header to format",
    show_default=True,
)
@click.option(
    "--row",
    default=1,
    help="The row of the header to format",
    show_default=True,
)
@click.option("--all", is_flag=True, help="Formats all columns in the row")
def format_header(xlsx_path: str, sheet_name: str, column: str, row: int, all: bool):
    workbook = load_workbook(xlsx_path)
    format(workbook, column, row, sheet_name, all)
    workbook.save(xlsx_path)


@main.command()
@click.argument("xlsx_path", type=click.Path(exists=True), required=True)
@click.argument("column_name", required=True)
def add_column(xlsx_path: str, column_name: str):
    workbook = load_workbook(xlsx_path)
    add_column_to_sheets(workbook, column_name, [])
    workbook.save(xlsx_path)


@main.command()
@click.argument("xlsx_path", type=click.Path(exists=True), required=True)
@click.argument("query", required=True)
@click.argument("result_column", required=True)
@click.option(
    "--sheet-name",
    help="The name of the sheet to format, if not provided will format all sheets",
)
@click.option(
    "--true-value",
    default="True",
    help="The value to set when the query is true",
    show_default=True,
)
@click.option(
    "--false-value",
    default="False",
    help="The value to set when the query is false",
    show_default=True,
)
def query(
    xlsx_path: str,
    query: str,
    result_column: str,
    sheet_name: str,
    true_value: str,
    false_value: str,
):
    workbook = load_workbook(xlsx_path)
    query_command(workbook, query, result_column, sheet_name, true_value, false_value)
    workbook.save(xlsx_path)


@main.command(
    help="Filters rows by values of a column and return data (or hyperlink) of another column"
)
@click.argument("xlsx_path", type=click.Path(exists=True), required=True)
@click.argument(
    "sheet_name",
    # help="Name of the sheet to search. If not given, all sheets are searched",
    required=True,
)
@click.argument(
    "filter_column",
    # help="Column in which a value is searched",
    required=True,
)
@click.argument(
    "filter_value",
    # help="Value to search for in the search column",
    required=True,
)
@click.argument(
    "column",
    #   help="Column to get the data from",
    required=True,
)
@click.option(
    "--url/--no-url",
    default=False,
    help="If given, returns the URL and not the text content of the resulting cell",
)
@click.option(
    "--return-all/--no-return-all",
    default=False,
    help="If given, returns data from all matching rows, not just the first",
)
@click.option(
    "--skip-rows",
    default=0,
    help="Number of header rows to skip",
)
def get_row(
    xlsx_path: str,
    sheet_name: str,
    filter_column: str,
    filter_value: str,
    column: str,
    url: bool,
    return_all: bool,
    skip_rows: int,
):
    workbook = load_workbook(xlsx_path, read_only=False)
    print_cell_value_from_filtered_row(
        workbook, sheet_name, filter_column, filter_value, column, url, return_all, skip_rows
    )


@main.group(no_args_is_help=True)
def evaluate():
    pass


evaluate.add_command(check_columns_command)


@main.command()
@click.argument("xlsx_path", type=click.Path(exists=True), required=True)
@click.argument("column_name", required=True)
def add_empty_column(xlsx_path: str, column_name: str):
    workbook = load_workbook(xlsx_path)
    add_column_to_sheets(workbook, column_name, [])
    workbook.save(xlsx_path)


if __name__ == "__main__":
    main()
